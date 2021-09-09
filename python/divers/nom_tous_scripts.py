from os import listdir
from os.path import isfile, join
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in cell.value.split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)

def create_excel(path_excel):
    wb = Workbook()
    sheet = wb.create_sheet("Scripts de curation")
    sheet.append(["Catégorie", "Script"])

    mypath = 'C:\\Users\\mboutrou\\Documents\\scripts\\curation'
    directories = [f for f in listdir(mypath) if not(isfile(join(mypath, f)))]

    n_ligne = 2
    for dir in directories:
        files = [f for f in listdir(mypath+'\\'+dir) if isfile(join(mypath+'\\'+dir, f))]

        files2 = sorted(files, key=lambda str: int(str.split('_')[0]))

        for file in files2 :
            sheet.append([dir, file])
            n_ligne += 1

        sheet.merge_cells(start_row=n_ligne-len(files2), start_column=1, end_row=n_ligne-1, end_column=1)

    redimension_cell_width(sheet)
    borders_cells(sheet)
    style_sheet(sheet)

    del wb["Sheet"]
    wb.save(str(path_excel))

path_excel = r"C:\Users\Public\Documents\liste_des_scripts.xlsx"
create_excel(path_excel)