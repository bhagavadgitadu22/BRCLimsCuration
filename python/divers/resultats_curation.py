#!/usr/bin/env python
import html
import re

from sqlalchemy import create_engine
from sqlalchemy.ext.automap import automap_base
from sqlalchemy.future import select
from sqlalchemy.orm import Session

import psycopg2

from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font


comptage_cles = []

def remove_curlystags(string=None):
    if string is not None and string != '':
        return html.unescape(re.sub(r'\</?[a-zA-Z ="0-9]+/?\>', '', string))
    else:
        return None

def _name_for_collection_relationship(base, local_cls, referred_cls, constraint):
    if constraint.name in comptage_cles:
        comptage_cles.append(constraint.name)
        return constraint.name+str(comptage_cles.count(constraint.name))
    else:
        comptage_cles.append(constraint.name)
        return constraint.name
    # if this didn't work, revert to the default behavior

def get_table_objects(engine):
    Base = automap_base()
    # reflect the tables
    Base.prepare(engine, reflect=True, name_for_collection_relationship=_name_for_collection_relationship)

    Souche = Base.classes.t_souche
    Dico = Base.classes.t_donneedico
    Collection = Base.classes.t_collection

    return {'souche': Souche, 'dico': Dico, 'collection': Collection}

def main():
    # initialisation
    engine = create_engine(f"postgresql://postgres:hercule1821@localhost:5432/brc_db")
    tables = get_table_objects(engine)

    engine_curated = create_engine(f"postgresql://postgres:hercule1821@localhost:5432/brc_db")
    tables_curated = get_table_objects(engine_curated)

    # on récupère tous les ids de souches puis on les parcourt un à un
    print("On vérifie que plus d'ids avant curation que après curation")
    
    liste_ids = liste_ids_souches(engine, tables)
    liste_ids_curated = liste_ids_souches(engine_curated, tables_curated)

    for xxx_id in liste_ids_curated:
        if xxx_id not in liste_ids:
            print("Un id de plus une fois la base curée ce n'est vraiment pas normal...")
            return None

    # on ne garde que souches de cip
    # en vérifiant d'abord que les autres n'ont pas été modifiées
    print("On récupère la liste des souches de la cip")

    collections_cip = []
    Collection = tables['collection']
    with Session(bind=engine, future=True) as session:
        statement = select(Collection).where(Collection.col_clg_id == 401)

        for row in session.execute(statement):
            collections_cip.append(row.t_collection.xxx_id)

    # on récupère arbres de taxo pour deux bases
    taxo = arbre_taxo("brc_db")
    taxo_curated = arbre_taxo("brc_db_curated")

    print("On parcourt les souches pour trouver les identifiants supprimés")

    ids_changes = []
    ids_deleted = []

    erreurs = []
    for xxx_id in liste_ids:
        if xxx_id in liste_ids_curated:
            souche = get_souche(xxx_id, engine, tables)
            souche_curated = get_souche(xxx_id, engine_curated, tables_curated)

            print(souche)
            print(souche_curated)
            print(compare(souche, souche_curated))
            
            # seul "problème", ça ne tient pas compte de différence éventuelle dans les parents à défaut des enfants pour taxonomie
            if souche = souche_curated:
                ids_changes.append(xxx_id)
            else:
                # checker si ça fait bien partie de la collection de la cip
                if souche.t_souche.sch_col_id not in collections_cip or souche_curated.t_souche.sch_col_id not in collections_cip:
                    print(souche)
                    print(souche_curated)
                    print("Erreur : une souche hors de la CIP a été modifié")
                    erreurs.append(souche.t_souche.sch_identifiant)
                    
        else:
            # on garde la trace des ids_supprimes, qui doivent correspondre à des ids 2xxxxx
            ids_deleted.append(xxx_id)

    print(str(len(erreurs)))
    print(erreurs)
    print(str(len(ids_changes)))
    return None
    
    print("On parcourt les "+str(len(ids_changes))+" souches avec des modifs pour trouver lesquelles ont été faites")
    differents_historiques, differents_lieux, differents_lieux_avec_date, differentes_pathos, differentes_taxonomies, differentes_temperatures = comparaison_souches(ids_changes, engine, tables, taxo, engine_curated, tables_curated, taxo_curated)
    
    print("On crée l'Excel final")
    create_excel(differents_historiques, differents_lieux, differents_lieux_avec_date, differentes_pathos, differentes_taxonomies, differentes_temperatures)

def liste_ids_souches(engine, tables):
    Souche = tables['souche']

    with Session(bind=engine, future=True) as session:
        statement = select(Souche).order_by(Souche.xxx_id)
        
        liste_ids = []
        for souche in session.execute(statement):
            xxx_id = souche.t_souche.xxx_id
            liste_ids.append(xxx_id)

    return liste_ids

def get_souche(xxx_id, engine, tables):
    Souche = tables['souche']

    with Session(bind=engine, future=True) as session:
        statement = select(Souche).where(Souche.xxx_id == xxx_id)
        souche = session.execute(statement).one()

        return souche

def arbre_taxo(nom_bdd):
    connection = psycopg2.connect(user="postgres",
                                  password="hercule1821",
                                  host="localhost",
                                  port="5432",
                                  database=nom_bdd)
    connection.autocommit = True

    # Create a cursor to perform database operations
    cursor = connection.cursor()

    # initialization
    cursor.execute(open("../curation/80_taxonomie/stats/parentele_de_toutes_taxonomies.sql ", "r").read())
    records = cursor.fetchall()

    paths_taxo = {}
    for record in records:
        paths_taxo[record[1]] = record[3]

    return paths_taxo

def comparaison_souches(all_ids, engine, tables, taxo, engine_curated, tables_curated, taxo_curated):
    differents_historiques = []
    differents_lieux = []
    differents_lieux_avec_date = []
    differentes_pathos = []
    differentes_taxonomies = []
    differentes_temperatures = []

    i = 0
    for xxx_id in all_ids:
        elmt = get_caracs_souche(xxx_id, engine, tables, taxo)
        elmt_curated = get_caracs_souche(xxx_id, engine_curated, tables_curated, taxo_curated)

        if elmt == elmt_curated:
            print(xxx_id+" a été changé mais pas dans les champs normaux ???")
            return None
        else:
            if elmt["sch_identifiant"] != elmt_curated["sch_identifiant"]:
                print("L'identifiant devrait être le même étrange ???")

            if elmt["sch_historique"] != elmt_curated["sch_historique"]:
                differents_historiques.append([elmt["sch_identifiant"], elmt["sch_historique"], elmt_curated["sch_historique"]])

            if elmt["dico_lieu"] != elmt_curated["dico_lieu"] or elmt["sch_lieu_precis"] != elmt_curated["sch_lieu_precis"]:
                if elmt["sch_dat_acquisition"] == elmt_curated["sch_dat_acquisition"]:
                    differents_lieux.append([elmt["sch_identifiant"], elmt["dico_lieu"], elmt_curated["dico_lieu"], elmt["sch_lieu_precis"], elmt_curated["sch_lieu_precis"]])
                else:
                    differents_lieux_avec_date.append([elmt["sch_identifiant"], elmt["dico_lieu"], elmt_curated["dico_lieu"], elmt["sch_lieu_precis"], elmt_curated["sch_lieu_precis"], elmt["sch_dat_acquisition"], elmt_curated["sch_dat_acquisition"]])

            if elmt["dico_patho_animal"] != elmt_curated["dico_patho_animal"]:
                differentes_pathos.append([elmt["sch_identifiant"], elmt["dico_patho_animal"], elmt_curated["dico_patho_animal"]])

            if elmt["sch_taxonomie"] != elmt_curated["sch_taxonomie"]:
                differentes_taxonomies.append([elmt["sch_identifiant"], elmt["sch_taxonomie"], elmt_curated["sch_taxonomie"]])

            if elmt["sch_temperature_incubation"] != elmt_curated["sch_temperature_incubation"]:
                differentes_temperatures.append([elmt["sch_identifiant"], elmt["sch_temperature_incubation"], elmt_curated["sch_temperature_incubation"]])

        i += 1
        if i%100 == 0:
            print(str(i)+" souches traitées")
    
    return differents_historiques, differents_lieux, differents_lieux_avec_date, differentes_pathos, differentes_taxonomies, differentes_temperatures

def get_caracs_souche(xxx_id, engine, tables, taxo):
    Souche = tables['souche']
    Dico = tables['dico']

    with Session(bind=engine, future=True) as session:
        statement = select(Souche).where(Souche.xxx_id == xxx_id)
        souche = session.execute(statement).one()

        xxx_id = souche.t_souche.xxx_id

        rows_lieu = select(Dico).where(Dico.xxx_id == souche.t_souche.sch_lieu)
        row_lieu = session.execute(rows_lieu).one()

        rows_patho = select(Dico).where(Dico.xxx_id == souche.t_souche.sch_patho_animal)
        row_patho = session.execute(rows_patho).one()

        elmt = {'sch_identifiant': souche.t_souche.sch_identifiant, 

                'sch_historique': souche.t_souche.sch_historique, 

                'dico_lieu': row_lieu.t_donneedico.don_lib, 
                'sch_lieu_precis': souche.t_souche.sch_lieu_precis, 
                'sch_dat_acquisition': souche.t_souche.sch_dat_acquisition, 

                'dico_patho_animal' : row_patho.t_donneedico.don_lib, 
                
                'sch_taxonomie': taxo[souche.t_souche.sch_taxonomie], 
                
                'sch_temperature_incubation': souche.t_souche.sch_temperature_incubation}

    return elmt

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in cell.value.split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def wrap_lines(sheet):
    for col in sheet.rows:
        col[0].alignment = Alignment(wrapText=True)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)
    
    wrap_lines(sheet)
    redimension_cell_width(sheet)
    borders_cells(sheet)

def create_sheet(wb, array, array_name, name):
    sheet = wb.create_sheet(array_name)

    sheet.append(["sch_identifiant", "old_"+name, "new_"+name])
    for element in array:
        sheet.append(element)
    
    style_sheet(sheet)

def create_excel(differents_historiques, differents_lieux, differents_lieux_avec_date, differentes_pathos, differentes_taxonomies, differentes_temperatures):
    wb = Workbook()

    create_sheet(wb, differents_historiques, "differents_historiques", "historique")

    sheet_lieu = wb.create_sheet("differents_lieux")
    sheet_lieu.append(["sch_identifiant", "old_lieu", "new_lieu", "old_lieu_precis", "new_lieu_precis"])
    for element in differents_lieux:
        sheet_lieu.append(element)
    style_sheet(sheet_lieu)

    sheet_lieu_date = wb.create_sheet("differents_lieux_avec_date")
    sheet_lieu_date.append(["sch_identifiant", "old_lieu", "new_lieu", "old_lieu_precis", "new_lieu_precis", "old_date", "new_date"])
    for element in differents_lieux_avec_date:
        sheet_lieu_date.append(element)
    style_sheet(sheet_lieu_date)

    create_sheet(wb, differentes_pathos, "differentes_pathos", "patho")
    create_sheet(wb, differentes_taxonomies, "differentes_taxonomies", "taxonomie")
    create_sheet(wb, differentes_temperatures, "differentes_temperatures", "temperature")

    del wb["Sheet"]
    wb.save(r"C:\Users\Public\Documents\bilan_global_curation.xlsx")

main()