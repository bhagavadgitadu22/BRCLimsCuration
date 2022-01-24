import os
import csv
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in cell.value.split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value*1.2

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font = Font(bold=True)

    redimension_cell_width(sheet)
    borders_cells(sheet)

#path = 'X:/crbtous/genomes_care'
path = 'X:/crbtous/genomes_care'
dir_list = os.listdir(path)

wb = Workbook()

# puis on lit les fichiers des différents dossiers pour voir si on a tout téléchargé
total_ids = 0
total_fastq = 0
old_total = 0
for genus in dir_list:
    if not(genus.endswith(".xlsx")) and genus != 'sratoolkit.2.11.3-win64' and genus != '.DS_Store' and genus != 'Staphylococcus2':
        print("")
        print(genus)

        sheet = wb.create_sheet(genus)

        # on s'occupe des colonnes du fichier de klebsiella
        cols = ["RM ID", "CARE ID", "SHORT WGS", "ERR/SRR ASSOCIATED", "FILENAME"]
        sheet.append(cols)

        # on récupère liste des errs qu'on voulait télécharger
        ids = open(path+'/'+genus+'/ids.csv', 'r', newline='', encoding='utf-8')
        csvreader = csv.reader(ids, delimiter=';')

        # on récupère liste de tous errs dont on a fasta pour ce génome
        local_files = os.listdir(path+'/'+genus)
        local_errs =  []

        n_ligne = 2
        used_files = []
        for row_csv in csvreader:
            number_fastq = 0
            for file in local_files:
                if not(file.endswith(".csv")) and not(file.endswith(".txt")):
                    err_downloaded = file.split(".")[0].split("_")[0]

                    if err_downloaded == row_csv[3]:
                        final_row = [elmt for elmt in row_csv]
                        final_row.append(file)
                        sheet.append(final_row)

                        n_ligne += 1
                        number_fastq += 1
                        used_files.append(file)

            if number_fastq == 0:
                print("fastq missing: "+str(row_csv[3]))

            if number_fastq > 1:
                for column in range(1, 5):
                    sheet.merge_cells(start_row=n_ligne-number_fastq, start_column=column, end_row=n_ligne-1, end_column=column)

            total_ids += 1
        print(total_ids-old_total)
        old_total = total_ids

        style_sheet(sheet)

        for file in local_files:
            if file not in used_files and not(file.endswith(".csv")) and not(file.endswith(".txt")):
                print("useless file: "+str(file))

        total_fastq += len(local_files)

del wb["Sheet"]
wb.save("X:/crbtous/genomes_care/final_ids.xlsx")

print("total_ids: "+str(total_ids))
print("total_fastq: "+str(total_fastq))
