import pandas as pd
import os
import shutil

from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in str(cell.value).split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)
    
    redimension_cell_width(sheet)
    borders_cells(sheet)

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

xls = pd.ExcelFile('../../transfert_p2m/50_transfert_infos_p2m.xlsx')

#pos_genome = 5
pos_genome = 2

# list_ids = []
# df = pd.read_excel(xls, 'genomes')
# df = df.reset_index(drop=True)  # make sure indexes pair with number of rows
# for index, row in df.iterrows():
#     if row[1].split('/')[pos_genome] not in list_ids:
#         list_ids.append(row[1].split('/')[pos_genome])

# sub_ids = list_ids[10:100]
# print(sub_ids)

sub_ids = ['109252T', '109585T', '111175', '111411T', '111486T', '111608T', '111725T', '111733T', '111750T', '111751T', '60.14', 'A231', '101049T', '103452T', '103549', '103857', '103869', '106155T', '111065', '111402', '111403', '111444T', '111497T', '111542T', '111543T', '111544T', '111546T', '64.31T', '105568', '111937', '78.25', '106788', '111830', '111844T', '111847T', '111848T', '105702', '105701', '111177T', 'A30', '105674', '102697', 'A31', '104583', '111179T', '111176T', '102722T', '102728', '103551', '103716', '103744T', '104221', '105921', '105922', '109891T', '110923', '111670', '111686', '111696T', '111715T', '111717T', '103212', '103570', '104340', '105891', '105960T', '106356T', '106358T', '107124T', '107448T', '107652', '108050T', '109356T', '109574T', '109743T', '109815T', '110183T', '111129', '111172', '111218', '111400T', '55.30', '75.1T', '81.57T', 'CRBIP17.18', '105662']

dirs_created = []
metafiles_moved = []
genomes_moved = []

df = pd.read_excel(xls, 'directories')
df = df.reset_index(drop=True)  # make sure indexes pair with number of rows
for index, row in df.iterrows():
    id = row[0].split('/')[pos_genome]

    if id in sub_ids:
        dir = row[0]
        if not(os.path.exists(dir)):
            os.mkdir(dir)
            dirs_created.append(dir)
        else:
            print('')
            print(dir)

df = pd.read_excel(xls, 'metafiles')
df = df.reset_index(drop=True)  # make sure indexes pair with number of rows
for index, row in df.iterrows():
    id = row[1].split('/')[pos_genome]

    if id in sub_ids:
        old_location = row[0]
        new_location = row[1]

        if os.path.exists(old_location):
            shutil.copyfile(old_location, new_location)
            temp = [old_location, new_location]
            metafiles_moved.append(temp)
        else:
            print('')
            print(old_location)
            print(new_location)

df = pd.read_excel(xls, 'genomes')
df = df.reset_index(drop=True)  # make sure indexes pair with number of rows
for index, row in df.iterrows():
    id = row[1].split('/')[pos_genome]

    if id in sub_ids:
        old_location = row[0]
        new_location = row[1]
        if os.path.exists(old_location):
            os.rename(old_location, new_location)
            temp = [old_location, new_location]
            genomes_moved.append(temp)
        else:
            print('')
            print(old_location)
            print(new_location)

wb = Workbook()

sheet = wb.create_sheet('directories')
del wb["Sheet"]
cols = ["dir_created"]
sheet.append(cols)
for line_dir in dirs_created:
    sheet.append([line_dir])
style_sheet(sheet)

sheet = wb.create_sheet('metafiles')
cols = ["old_location", "new_location"]
sheet.append(cols)
for line_meta in metafiles_moved:
    sheet.append(line_meta)
style_sheet(sheet)

sheet = wb.create_sheet('genomes')
cols = ["old_location", "new_location"]
sheet.append(cols)
for line_genome in genomes_moved:
    sheet.append(line_genome)
style_sheet(sheet)

wb.save("../../transfert_p2m/60_bilan_transfert_p2m.xlsx")
