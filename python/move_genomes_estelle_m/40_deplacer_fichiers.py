import pandas as pd
import os
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in str(cell.value).split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)
    
    redimension_cell_width(sheet)
    borders_cells(sheet)

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

xls_milieux = pd.ExcelFile('../../output/all_infos_p2m_crossed.xlsx')
df = pd.read_excel(xls_milieux, 'genomes_crossed')

#local_path = '/mnt/gaia/cip'
#pos_p2m = 6
local_path = 'X:'
pos_p2m = 3
path = local_path+'/CONTROLE_SOUCHE'
dir_list = os.listdir(path)

dirs_created = []
dirs_existent = []

log_transferts = []

compet = 0

print("start of search in genomes")
df = df.reset_index(drop=True)  # make sure indexes pair with number of rows
for index, row in df.iterrows():
    # si on a confronté avec succès l'id ET le lot au fichier EM
    if bool(row[6]) and bool(row[7]) and row[2] != '':
        id_cip = row[0].replace("CIP", '').replace("CRBIP", '').replace("_", '').replace("-", '').replace(".", '').replace(" ", '').replace("T", '')
        lot_cip = str(row[1]).replace("_", '').replace("-", '').replace(".", '').replace(" ", '')
        fasta_cip = row[2]

        fileToTransfer = row[3].split('/')[-1]
        compte = row[4]

        row_transfert = [row[3], '']

        # si j'ai trouvé un dossier associé à l'id en cours
        # alors je peux checker si le fichier qui m'occupe est déjà transféré ou non
        # si transfert est vrai je ne fais rien sinon je peux procéder au transfert du fichier
        dir_id = ''
        dir_lot = ''
        dir_fasta = ''
        for file_id in dir_list:
            short_file_id = file_id.replace("CIP", '').replace("CRBIP", '').replace("_", '').replace("-", '').replace(".", '').replace(" ", '').replace("T", '')

            if short_file_id == id_cip and os.path.isdir(os.path.join(path, file_id)):
                dir_id = path+'/'+file_id

                # on regarde s'il y a un dossier pour le lot à l'intérieur
                id_list = os.listdir(dir_id)
                for lot_dir in id_list:
                    if lot_dir == file_id+'-'+lot_cip and os.path.isdir(os.path.join(dir_id, lot_dir)):
                        dir_lot = dir_id+'/'+lot_dir

                        # on regarde s'il y a un dossier pour le type de fasta à l'intérieur
                        lot_list = os.listdir(dir_lot)
                        for fasta_dir in lot_list:
                            if fasta_dir == fasta_cip and os.path.isdir(os.path.join(dir_lot, fasta_dir)):
                                dir_fasta = dir_lot+'/'+fasta_dir

                                elmts_genomes = os.listdir(dir_fasta)
                                for elmt_genome in elmts_genomes:
                                    if elmt_genome == fileToTransfer:
                                        # on compte les fichiers identiques nous attendant dans la cible auquel cas on n'agit pas
                                        compte += 1

        # je crée les dossiers qu'il manque à ce stade
        if dir_id == '':
            dir_id = path+'/'+row[0].replace("CIP", '').replace("-", '.')
            #os.mkdir(dir_id)
            if dir_id not in dirs_created:
                dirs_created.append(dir_id)
        elif dir_id not in dirs_existent:
            dirs_existent.append(dir_id)

        if dir_lot == '':
            dir_lot = dir_id+'/'+row[0].replace("CIP", '').replace("-", '.')+'-'+row[1].replace("-", '.')
            #os.mkdir(dir_lot)
            if dir_lot not in dirs_created:
                dirs_created.append(dir_lot)
        elif dir_lot not in dirs_existent:
            dirs_existent.append(dir_lot)

        if dir_fasta == '':
            dir_fasta = dir_lot+'/'+row[2]
            #os.mkdir(dir_fasta)
            if dir_fasta not in dirs_created:
                dirs_created.append(dir_fasta)
        elif dir_fasta not in dirs_existent:
            dirs_existent.append(dir_fasta)

        # si fichier dans un dir_fasta le copier coller dans dir_fasta
        final_location = dir_fasta+'/'+fileToTransfer
        if compte != 1:
            final_location = dir_fasta+'/'+str(compte)+'_'+fileToTransfer
        row_transfert[1] = final_location

        log_transferts.append(row_transfert)

        compet += 1
        if compet % 1000 == 0:
            print(compet)

    # if compet == 2500:
    #     break

p2m_used = []
for line_log in log_transferts:
    if line_log[0].split('/')[pos_p2m] not in p2m_used:
        p2m_used.append(line_log[0].split('/')[pos_p2m])

df = pd.read_excel(xls_milieux, 'metafiles')

print("start of search in metafiles")
log_metafiles = []
unfound = []
df = df.reset_index(drop=True)  # make sure indexes pair with number of rows
for index, row in df.iterrows():
    n_p2m = row[1].split('/')[pos_p2m]
    filename = row[1].split('/')[-1]

    if n_p2m in p2m_used:
        if filename in ['FASTA.info.txt', 'SampleSheet.csv', 'README.txt', 'FASTQ.info.txt', 'SampleSheet.info.csv', 'README.pdf']:
            compte_meta = 1
            for elmt_meta in log_metafiles:
                if path+'/'+row[0].replace("CIP", '').replace("-", '.')+'/'+filename == elmt_meta[1]:
                    compte_meta += 1

            row_metafile = [row[1], path+'/'+row[0].replace("CIP", '').replace("-", '.')+'/'+filename]
            if compte_meta != 1:
                row_metafile = [row[1], path+'/'+row[0].replace("CIP", '').replace("-", '.')+'/'+str(compte_meta)+'_'+filename]

            log_metafiles.append(row_metafile)
        else:
            unfound.append(row[1])

# création de l'excel avec toutes les infos
wb = Workbook()

sheet = wb.create_sheet('directories')
del wb["Sheet"]
cols = ["dir_created"]
sheet.append(cols)
for line_dir in dirs_created:
    sheet.append([line_dir])
style_sheet(sheet)

sheet = wb.create_sheet('dirs_from_before')
cols = ["dir_existent"]
sheet.append(cols)
for line_dir in dirs_existent:
    sheet.append([line_dir])
style_sheet(sheet)

sheet = wb.create_sheet('genomes')
cols = ["old_location", "new_location"]
sheet.append(cols)
for line_log in log_transferts:
    sheet.append(line_log)
style_sheet(sheet)

sheet = wb.create_sheet('metafiles')
cols = ["old_location", "new_location"]
sheet.append(cols)
for line_meta in log_metafiles:
    sheet.append(line_meta)
style_sheet(sheet)

sheet = wb.create_sheet('weird_metafile')
cols = ["file"]
sheet.append(cols)
for line_weird in unfound:
    sheet.append([line_weird])
style_sheet(sheet)

wb.save("../../output/transfert_infos_p2m.xlsx")
