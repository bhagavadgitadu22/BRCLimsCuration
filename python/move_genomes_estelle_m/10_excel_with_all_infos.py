import os
import re
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def test_match(pattern, file):
    id_cip = ''
    match = re.match(pattern, file)
    if match:
        id_cip = match.group(0)
        if '105521' in id_cip:
            print('')
            print(pattern)
            print(file)
            print(id_cip)
    return id_cip

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in str(cell.value).split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)
    
    redimension_cell_width(sheet)
    borders_cells(sheet)

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#local_path = '/mnt/gaia/cip'
local_path = 'X:'
path = local_path+'/SEQUENCAGETOTAL/FICHIERRESULTATSSEQUENCES'
dir_list = os.listdir(path)

pattern1 = '^(CIP)?-?[0-9]{6}T?'
pattern2 = '^(CIP)?[0-9]{1,2}(-|.)[0-9]+T?'
pattern3 = '^CRBIP[0-9]{1,3}-[0-9]+T?'
pattern4 = '^CIPA[0-9]{1,3}T?'

count = 0

# liste des dossiers à éviter car on préviligie copie
dossiers_copie = []
for file in dir_list:
    # on traite chaque dossier de p2m
    if os.path.isdir(os.path.join(path, file)) and file[0].isdigit() and "copie" in file.lower():
        dossiers_copie.append(file.replace("-Copie", ""))

print(dossiers_copie)

# file is the number p2m
ids_cip = {}
ids_cip_extra = {}
fake_ids = []
for file in dir_list:
    # on traite chaque dossier de p2m
    if os.path.isdir(os.path.join(path, file)) and file[0].isdigit() and file not in dossiers_copie:
        path2 = path+'/'+file
        dir_list2 = os.listdir(path2)

        # création du dico avec toutes les infos utiles
        extra_files = []
        ids_concerned = []

        # f2 is the type of file : fasta, fasta.flt...
        for f2 in dir_list2:
            path3 = path2+'/'+f2

            if not(os.path.isdir(path3)):
                if 'DS_Store' not in f2:
                    extra_files.append(path3)
            else:
                dir_list3 = os.listdir(path3)

                # f3 are the names of the genomes' files
                for f3 in dir_list3:
                    id_found = True
                    if test_match(pattern1, f3) != '':
                        id_cip = test_match(pattern1, f3)
                    elif test_match(pattern2, f3) != '':
                        id_cip = test_match(pattern2, f3)
                    elif test_match(pattern3, f3) != '':
                        id_cip = test_match(pattern3, f3)
                    elif test_match(pattern4, f3) != '':
                        id_cip = test_match(pattern4, f3)
                    else:
                        if 'DS_Store' not in f3:
                            fake_ids.append(path3+'/'+f3)
                        id_found = False

                    # si je n'ai pas trouvé l'id ça ne sert à rien de s'acharner sur ce fichier
                    if id_found:
                        if id_cip not in ids_cip:
                            ids_cip[id_cip] = {}
                        if id_cip not in ids_concerned:
                            ids_concerned.append(id_cip)

                        str_lot = f3.replace(id_cip, '')

                        lot_cip = ''
                        if str_lot not in [''] and str_lot[0] in ['_', '-']:
                            lot_cip = str_lot[1:]
                            if '_' in str_lot[1:]:
                                lot_cip = str_lot[1:].split('_')[0]
                            elif '.' in str_lot[1:]:
                                lot_cip = str_lot[1:].split('.')[0]

                            if lot_cip != '' and lot_cip[0] == 'S':
                                lot_cip = ''

                        if lot_cip not in ids_cip[id_cip]:
                            ids_cip[id_cip][lot_cip] = {}
                            
                        fasta_dir = path3.split('/')[-1]
                        if fasta_dir not in ids_cip[id_cip][lot_cip]:
                            ids_cip[id_cip][lot_cip][fasta_dir] = {}

                        fasta_file = path3+'/'+f3
                        if fasta_file not in ids_cip[id_cip][lot_cip][fasta_dir]:
                            ids_cip[id_cip][lot_cip][fasta_dir][fasta_file] = 0

                            # on tient le compte du nombre de fois où chaque file est utilisé
                            for elmt_genome in ids_cip[id_cip][lot_cip][fasta_dir]:
                                if elmt_genome.split('/')[-1] == f3:
                                    ids_cip[id_cip][lot_cip][fasta_dir][fasta_file] += 1
        
        # il peut y avoir plusieurs fois un des 4 fichiers de base si plusieurs lots d'une même souche ont été séquencés
        for id_c in ids_concerned:
            for extra_f in extra_files:
                if id_c not in ids_cip_extra:
                    ids_cip_extra[id_c] = {}

                ids_cip_extra[id_c][extra_f] = 0
                
                # on tient le compte du nombre de fois où chaque file est utilisé
                for elmt_genome in ids_cip_extra[id_c]:
                    if elmt_genome.split('/')[-1] == extra_f.split('/')[-1]:
                        ids_cip_extra[id_c][extra_f] += 1

    count += 1
    if count % 100 == 0:
        print(count)

    # if "copie" in file.lower():
    #     print(file)

# création de l'excel avec toutes les infos
wb = Workbook()
sheet = wb.create_sheet('genomes')
del wb["Sheet"]

cols = ["id", "lot", "dir", "path", "compte"]
sheet.append(cols)
for id in ids_cip:
    for lot in ids_cip[id]:
        for dir in ids_cip[id][lot]:
            for f in ids_cip[id][lot][dir]:
                row = [id, lot, dir, f, ids_cip[id][lot][dir][f]]
                sheet.append(row)
style_sheet(sheet)

sheet2 = wb.create_sheet('metafiles')
cols = ["id", "path", "compte"]
sheet2.append(cols)
used = []
for id in ids_cip_extra:
    for f in ids_cip_extra[id]:
        if 'readme' in f.lower():
            if id not in used:
                first_use = id
                used.append(first_use)
                row = [id, f, ids_cip_extra[id][f]]
                sheet2.append(row)
            else:
                print([id, f])
        else:
            row = [id, f, ids_cip_extra[id][f]]
            sheet2.append(row)
style_sheet(sheet2)

wb.save("../../output/all_infos_p2m.xlsx")

# et de celui avec les fichiers dont je n'ai pas trouvé l'id
wb = Workbook()
sheet = wb.create_sheet('missed')
del wb["Sheet"]
cols = ["fake_ids"]
sheet.append(cols)
for id in fake_ids:
    sheet.append([id])
style_sheet(sheet)
wb.save("../../output/fake_ids_p2m.xlsx")
