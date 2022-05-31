import urllib.request
import os
import pandas as pd
from os.path import isfile, join
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in cell.value.split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+2

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font = Font(bold=True)

    header2 = list(sheet.rows)[1]
    for cell in header2:
        cell.font = Font(bold=True)

    for col in sheet.rows:
        cell = col[0]
        cell.font = Font(bold=True)

def style_sheet2(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font = Font(bold=True)

    for col in sheet.rows:
        cell = col[0]
        cell.font = Font(bold=True)

def merge_headers(sheet, header):
    value = header[1]
    count = 0
    for i in range(1, len(header)):
        if header[i] != value:
            sheet.merge_cells(start_row=1, start_column=i-count+1, end_row=1, end_column=i)
            value = header[i]
            count = 1
        else:
            count += 1

lines = []
link = "https://bitbucket.org/genomicepidemiology/resfinder_db/raw/fa32d9a3cf0c12ec70ca4e90c45c0d590ee810bd/phenotypes.txt"
with urllib.request.urlopen(link) as url:
    s = str(url.read())
    # I'm guessing this would output the html source code ?
    genes_rb = []
    classes_rb = []
    for elmt in s.split('\\n'):
        if '\\t' in elmt:
            genes_rb.append(elmt.split('\\t')[0].split('_')[0])
            classes_rb.append(elmt.split('\\t')[1])
        else:
            print(elmt)

genes_rb.append('mcr-1')
classes_rb.append('Polymyxin')

if len(genes_rb) != len(classes_rb):
    print("trouble!")

# local_path = 'C:/Users/mboutrou/Documents/'
local_path = '/mnt/gaia/my_home/'
path = local_path+'TR _AMR_database_pour_CARE'
dir_list = os.listdir(path)

genus_par_classe = {}
names_classes = {}
genes_par_classes_par_genus = {}
genes_par_classes = {}
for f in dir_list:
    genus = f.split('.')[0]
    print(genus)

    genes_genus = {}
    genes_par_classes_par_genus[genus] = {}

    xlsx = pd.ExcelFile(path+'/'+f)
    sheet_names = xlsx.sheet_names  # see all sheet names

    # je lis les feuilles de l'excel en cours une à une
    for sheet_name in sheet_names:
        sheet = pd.read_excel(xlsx, sheet_name)
        
        # on récupère les numéros des 3 colonnes d'intérêt
        col_care = -1
        for col in range(len(sheet.columns)):
            if sheet.columns[col] == 'Genotype':
                col_care = col
        
        if col_care != -1:
            # pour chaque ligne de données je récupère l'identifiant intéressant (care ou rm) et le numéro wgs
            for i in range(1, len(sheet)):
                care_id = sheet.iloc[i, col_care]
                genes = care_id.split(', ')

                for gene in genes:
                    gene = gene.replace('oqx', 'Oqx')
                    gene = gene.replace('strA', 'aph(3\'\')-Ib')

                    if gene in genes_rb:
                        idx = genes_rb.index(gene)
                        classe = classes_rb[idx]
                        if classe not in genes_genus:
                            genes_genus[classe] = {}

                        if gene not in genes_genus[classe]:
                            genes_genus[classe][gene] = 0
                        genes_genus[classe][gene] += 1

                        if classe not in names_classes:
                            names_classes[classe] = []
                        if gene not in names_classes[classe]:
                            names_classes[classe].append(gene)

                        if classe not in genes_par_classes_par_genus[genus]:
                            genes_par_classes_par_genus[genus][classe] = []
                        if gene not in genes_par_classes_par_genus[genus][classe]:
                            genes_par_classes_par_genus[genus][classe].append(gene)
                        if classe not in genes_par_classes:
                            genes_par_classes[classe] = []
                        if gene not in genes_par_classes[classe]:
                            genes_par_classes[classe].append(gene)


    genus_par_classe[genus] = genes_genus

# création de l'excel
wb = Workbook()

# feuille1
header = [""]
header2 = [""]
for n_c in names_classes:
    for g_c in names_classes[n_c]:
        header.append(n_c)
        header2.append(g_c)

rows = []
for genus in genus_par_classe:
    row = [genus]

    for i in range(1, len(header)):
        c_i = header[i]
        g_i = header2[i]

        if c_i in genus_par_classe[genus] and g_i in genus_par_classe[genus][c_i]:
            row.append(str(genus_par_classe[genus][c_i][g_i]))
        else:
            row.append(str(0))

    rows.append(row)

sheet = wb.create_sheet("Resistance_genes_used")
sheet.append(header)
sheet.append(header2)
for row in rows:
    sheet.append(row)

redimension_cell_width(sheet)
borders_cells(sheet)
style_sheet(sheet)
merge_headers(sheet, header)

# feuille2
header3 = list(set(header))
header3.sort()

rows = []
for genus in genes_par_classes_par_genus:
    row = [genus]
    sum = 0

    for i in range(1, len(header3)):
        c_i = header3[i]

        if c_i in genes_par_classes_par_genus[genus]:
            row.append(str(len(genes_par_classes_par_genus[genus][c_i])))
            sum += len(genes_par_classes_par_genus[genus][c_i])
        else:
            row.append(str(0))

    row.append(str(sum))
    rows.append(row)

row_final = ["Total par classe"]
sum_final = 0
for i in range(1, len(header3)):
    c_i = header3[i]

    if c_i in genes_par_classes:
        row_final.append(str(len(genes_par_classes[c_i])))
        sum_final += len(genes_par_classes[c_i])
    else:
        row_final.append(str(0))

row_final.append(str(sum_final))
rows.append(row_final)

sheet = wb.create_sheet("Statistics")
header3.append("Total par genre")
sheet.append(header3)
for row in rows:
    sheet.append(row)

total_resfinder = ["Total ResFinder"]
sum_res = 0
for i in range(1, len(header3)-1):
    c_f = header3[i]
    total_resfinder.append(str(classes_rb.count(c_f)))
    sum_res += classes_rb.count(c_f)
total_resfinder.append(str(sum_res))
sheet.append(total_resfinder) 

redimension_cell_width(sheet)
borders_cells(sheet)
style_sheet2(sheet)

# sauvegarde de l'excel
del wb["Sheet"]
path_excel = local_path+'bilan_genes.xlsx'
wb.save(str(path_excel))
