import urllib.request
import os
import pandas as pd
from os.path import isfile, join
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

### 0 - functions to style the Excel file
def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = len(str(cell.value))
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+2

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font = Font(bold=True)

    header2 = list(sheet.rows)[1]
    for cell in header2:
        cell.font = Font(bold=True)

    for col in sheet.rows:
        cell = col[0]
        cell.font = Font(bold=True)

def style_sheet2(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font = Font(bold=True)

    for col in sheet.rows:
        cell = col[0]
        cell.font = Font(bold=True)

def merge_headers(sheet, header):
    value = header[1]
    count = 0
    for i in range(1, len(header)):
        if header[i] != value:
            sheet.merge_cells(start_row=1, start_column=i-count+1, end_row=1, end_column=i)
            value = header[i]
            count = 1
        else:
            count += 1


### I - we read the phenotypes.txt file of the resfinder database from the internet
lines = []
link = "https://bitbucket.org/genomicepidemiology/resfinder_db/raw/fa32d9a3cf0c12ec70ca4e90c45c0d590ee810bd/phenotypes.txt"
with urllib.request.urlopen(link) as url:
    s = str(url.read())
    
    # then we analyse it to get the genes and the classes associated to each gene
    genes_rb = []
    classes_rb = []
    for elmt in s.split('\\n'):
        if '\\t' in elmt:
            genes_rb.append(elmt.split('\\t')[0].split('_')[0])
            classes_rb.append(elmt.split('\\t')[1])
        else:
            print(elmt)

genes_rb.append('mcr-1')
classes_rb.append('Polymyxin')

# background check: we must have as many genes as groups of classes
if len(genes_rb) != len(classes_rb):
    print("trouble!")


### II - then we open the CARE files with the results of the genotypes obtained in resfinder
local_path = 'C:/Users/mboutrou/Documents/'
#local_path = '/mnt/gaia/my_home/'
path = local_path+'TR _AMR_database_pour_CARE'
dir_list = os.listdir(path)

genus_par_classe = {}
names_classes = {}
genes_par_classes_par_genus = {}
genes_par_classes = {}
# there is one file for each genus
for f in dir_list:
    genus = f.split('.')[0]
    print(genus)

    genes_genus = {}
    genes_par_classes_par_genus[genus] = {}

    xlsx = pd.ExcelFile(path+'/'+f)
    sheet_names = xlsx.sheet_names  # see all sheet names

    # I try to locate the genotype column located in one of the sheets
    for sheet_name in sheet_names:
        sheet = pd.read_excel(xlsx, sheet_name)
        
        col_care = -1
        for col in range(len(sheet.columns)):
            if sheet.columns[col] == 'Genotype':
                col_care = col
        
        if col_care != -1:
            for i in range(1, len(sheet)):
                # then for each line I explode the genotypes cell to get the different genes listed
                care_id = sheet.iloc[i, col_care]
                genes = care_id.split(', ')

                # and I add one occurence of this gene on the list of genes for this genus (genes_genus)
                # and also one occurence for the corresponding class of genes on the list of classes for this genus
                for gene in genes:
                    gene = gene.replace('oqx', 'Oqx')
                    gene = gene.replace('strA', 'aph(3\'\')-Ib')

                    if gene in genes_rb:
                        idx = genes_rb.index(gene)
                        classe = classes_rb[idx]
                        if classe not in genes_genus:
                            genes_genus[classe] = {}

                        if gene not in genes_genus[classe]:
                            genes_genus[classe][gene] = 0
                        genes_genus[classe][gene] += 1

                        if classe not in names_classes:
                            names_classes[classe] = []
                        if gene not in names_classes[classe]:
                            names_classes[classe].append(gene)

                        if classe not in genes_par_classes_par_genus[genus]:
                            genes_par_classes_par_genus[genus][classe] = []
                        if gene not in genes_par_classes_par_genus[genus][classe]:
                            genes_par_classes_par_genus[genus][classe].append(gene)
                        if classe not in genes_par_classes:
                            genes_par_classes[classe] = []
                        if gene not in genes_par_classes[classe]:
                            genes_par_classes[classe].append(gene)

    genus_par_classe[genus] = genes_genus


### III - then I can start creating the final Excel
wb = Workbook()

# we create the first sheet
# the different genes are the headline and the different genus are the headrow
# each cell contain the number of occurence for the genus at the beginning of the row and the gene at the top of the column
header = [""]
header2 = [""]
for n_c in names_classes:
    for g_c in names_classes[n_c]:
        header.append(n_c)
        header2.append(g_c)

rows = []
for genus in genus_par_classe:
    row = [genus]

    for i in range(1, len(header)):
        c_i = header[i]
        g_i = header2[i]

        if c_i in genus_par_classe[genus] and g_i in genus_par_classe[genus][c_i]:
            row.append(genus_par_classe[genus][c_i][g_i])
        else:
            row.append(0)

    rows.append(row)

sheet = wb.create_sheet("Resistance_genes_used")
sheet.append(header)
sheet.append(header2)
for row in rows:
    sheet.append(row)

# and then we style the sheet
redimension_cell_width(sheet)
borders_cells(sheet)
style_sheet(sheet)
merge_headers(sheet, header)

# we create the second sheet
# it's the same thing than the first sheet but with the classes instead of the genes as the headline
header3 = list(set(header))
header3.sort()

rows = []
for genus in genes_par_classes_par_genus:
    row = [genus]
    sum = 0

    for i in range(1, len(header3)):
        c_i = header3[i]

        if c_i in genes_par_classes_par_genus[genus]:
            row.append(len(genes_par_classes_par_genus[genus][c_i]))
            sum += len(genes_par_classes_par_genus[genus][c_i])
        else:
            row.append(0)

    row.append(sum)
    rows.append(row)

row_final = ["All genera combined"]
sum_final = 0
for i in range(1, len(header3)):
    c_i = header3[i]

    if c_i in genes_par_classes:
        row_final.append(len(genes_par_classes[c_i]))
        sum_final += len(genes_par_classes[c_i])
    else:
        row_final.append(0)

row_final.append(sum_final)
rows.append(row_final)

sheet = wb.create_sheet("Statistics")
header3.append("All classes combined")
sheet.append(header3)
for row in rows:
    sheet.append(row)

total_resfinder = ["Total ResFinder"]
sum_res = 0
for i in range(1, len(header3)-1):
    c_f = header3[i]
    total_resfinder.append(classes_rb.count(c_f))
    sum_res += classes_rb.count(c_f)
total_resfinder.append(sum_res)
sheet.append(total_resfinder) 

# we style the sheet
redimension_cell_width(sheet)
borders_cells(sheet)
style_sheet2(sheet)

# then finally we save the Excel
del wb["Sheet"]
path_excel = local_path+'bilan_genes.xlsx'
wb.save(str(path_excel))
