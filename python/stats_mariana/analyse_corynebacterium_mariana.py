from openpyxl.workbook.workbook import Workbook
import psycopg2
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in cell.value.split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def wrap_lines(sheet):
    for col in sheet.rows:
        col[0].alignment = Alignment(wrapText=True)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)
    
    wrap_lines(sheet)
    redimension_cell_width(sheet)
    borders_cells(sheet)

def sheet_error(wb, cursor, file, name):
    cursor.execute(open("../analyse/corynebacterium/"+file, "r").read())
    records = cursor.fetchall()

    sheet = wb.create_sheet(name)
    sheet.append(["id" "souche", ""])

    n_ligne = 2
    for record in records:
        print(record[2])
        row = []

        # pour les sources d'erreur
        if isinstance(record[2], str):
            row.append(record[2])
        else :
            chaine = ""
            for erreur in record[2]:
                if chaine != "":
                    chaine += "\n- "
                chaine += str(erreur)
            row.append(chaine)
        
        # pour les noms de milieux et les souches qui en relèvent
        for nom in record[1]:
            for key in nom.keys():
                sheet.append(row+[str(key)]+[str(nom[key])])
            
            n_ligne += 1

        sheet.merge_cells(start_row=n_ligne-len(record[1]), start_column=1, end_row=n_ligne-1, end_column=1)

    style_sheet(sheet)

def create_excel(path):
    # on récupère la liste des identifiants valides
    connection = psycopg2.connect(user="postgres",
                                  password="hercule1821",
                                  host="localhost",
                                  port="5432",
                                  database="brc_db")
    connection.autocommit = True

    # Create a cursor to perform database operations
    cursor = connection.cursor()

    wb = Workbook()
    sheet_error(wb, cursor, "corynebacterium_antibios_mariana.sql", "recettes_eclatees")

    del wb["Sheet"]
    wb.save(str(path))

create_excel(r"C:\Users\Public\Documents\analyse_antibios_mariana.xlsx")