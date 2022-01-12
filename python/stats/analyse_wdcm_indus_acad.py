from openpyxl.workbook.workbook import Workbook
import psycopg2
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in str(cell.value).split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")

    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def wrap_lines(sheet):
    for col in sheet.rows:
        col[0].alignment = Alignment(wrapText=True)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font = Font(bold=True)
    
    wrap_lines(sheet)
    #redimension_cell_width(sheet)
    borders_cells(sheet)

def create_sheet(wb, lignes, annee, champs, acad_vs_indus):
    sheet = wb.create_sheet(str(annee))
    sheet.append(champs)

    print(annee)

    if acad_vs_indus:
        lignes_acad = []
        lignes_indus = []

        for ligne in lignes:
            if ligne[2] == "académique":
                lignes_acad.append(ligne)
            else:
                lignes_indus.append(ligne)
    
        n_lignes = max(len(lignes_acad), len(lignes_indus))
        lignes = [[]]*n_lignes

        i = 0
        while i < n_lignes:
            l = [' ']*9
            if i < len(lignes_acad):
                for j in range(4):
                    l[j] = lignes_acad[i][j]    
            if i < len(lignes_indus):  
                for j in range(4):  
                    l[5+j] = lignes_indus[i][j]
            
            lignes[i] = l
            i += 1

    for ligne in lignes:
        sheet.append(ligne)

    style_sheet(sheet)

def gestion_file(wb, records, champs, bool):
    lignes = []
    annee = 0

    for record in records:
        if record[0] != annee:
            if len(lignes) != 0:
                create_sheet(wb, lignes, annee, champs, bool)

            annee = record[0]
            lignes = [record]
        else:
            lignes.append(record)

    create_sheet(wb, lignes, annee, champs, bool)

def create_excel():
    # on récupère la liste des identifiants valides
    connection = psycopg2.connect(user="postgres",
                                  password="hercule1821",
                                  host="localhost",
                                  port="5432",
                                  database="brc_db")
    connection.autocommit = True

    # Create a cursor to perform database operations
    cursor = connection.cursor()

    # indus vs acad
    wb = Workbook()

    cursor.execute(open("../analyse/analyse_provenance_commandes/type_commande_par_annee.sql", "r").read())
    records = cursor.fetchall()

    gestion_file(wb, records, ["Année", "Institution", "Statut", "Quantité", "", "Année", "Institution", "Statut", "Quantité"], True)

    del wb["Sheet"]
    wb.save(r"C:\Users\Public\Documents\analyse_indus_vs_acad.xlsx")

    # wdcm
    wb = Workbook()

    cursor.execute(open("../analyse/analyse_provenance_commandes/wdcm_par_annee.sql", "r").read())
    records = cursor.fetchall()

    gestion_file(wb, records, ["Année", "Numéro WDCM", "Souches CIP", "Quantité"], False)

    del wb["Sheet"]
    wb.save(r"C:\Users\Public\Documents\analyse_wdcm.xlsx")

create_excel()
