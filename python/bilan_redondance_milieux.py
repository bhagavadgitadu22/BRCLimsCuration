from openpyxl.workbook.workbook import Workbook
import psycopg2
from openpyxl.styles import Border, Side, Alignment, Font

def redimension_cell_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                line_max = max([len(str(elmt)) for elmt in cell.value.split('\n')])
                max_ = max((dims.get(cell.column_letter, 0), line_max))
                dims[cell.column_letter] = max_
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def borders_cells(sheet):
    thin = Side(border_style="thin", color="000000")
    
    for col in sheet.rows:
        for cell in col:
            if cell.value:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def wrap_lines(sheet):
    for col in sheet.rows:
        col[0].alignment = Alignment(wrapText=True)

def style_sheet(sheet):
    header = list(sheet.rows)[0]
    for cell in header:
        cell.font  = Font(bold=True)
    borders_cells(sheet)
    wrap_lines(sheet)
    redimension_cell_width(sheet)

def designations_doublees(wb, cursor):
    cursor.execute(open("../analyse_milieux_culture/0_doublons_mil_designation_en.sql", "r").read())
    records = cursor.fetchall()

    sheet = wb.create_sheet("Désignations doublées")
    sheet.append(["Nom du milieu dupliqué", "IDs BDD"])

    for record in records:
        row = [record[1]]
        for id in record[0]:
            row.append(str(id))
        sheet.append(row)

    style_sheet(sheet)

def noms_pas_conformes(wb, cursor):
    cursor.execute(open("../analyse_milieux_culture/10_pas_medium_num_blabla.sql", "r").read())
    records = cursor.fetchall()

    sheet = wb.create_sheet("Noms pas conformes")
    sheet.append(["Nom du milieu pas conforme", "ID BDD"])

    for record in records:
        row = [record[1]]
        row.append(str(record[0]))
        sheet.append(row)

    style_sheet(sheet)

def commentaires_delateurs(wb, cursor):
    cursor.execute(open("../analyse_milieux_culture/15_milieu_double_mis_dans_commentaire.sql", "r").read())
    records = cursor.fetchall()

    sheet = wb.create_sheet("Commentaires délateurs")
    sheet.append(["Nom du milieu", "Commentaire l'incriminant", "ID BDD"])

    for record in records:
        row = [record[1]]
        row.append(record[2])
        row.append(str(record[0]))
        sheet.append(row)

    style_sheet(sheet)

def sheet_error(wb, cursor, file, name):
    cursor.execute(open("../analyse_milieux_culture/"+file, "r").read())
    records = cursor.fetchall()

    sheet = wb.create_sheet(name)
    sheet.append(["Erreur", "Nom des milieux"])

    n_ligne = 2
    for record in records:
        row = []
        if isinstance(record[2], str):
            row.append(record[2])
        else :
            chaine = ""
            for erreur in record[2]:
                if chaine != "":
                    chaine += "\n- "
                chaine += str(erreur)
            row.append(chaine)
        
        for nom in record[1]:
            sheet.append(row+[str(nom)])
            n_ligne += 1

        sheet.merge_cells(start_row=n_ligne-len(record[1]), start_column=1, end_row=n_ligne-1, end_column=1)

    style_sheet(sheet)

def create_excel(path):
    # on récupère la liste des identifiants valides
    connection = psycopg2.connect(user="postgres",
                                  password="hercule1821",
                                  host="localhost",
                                  port="5432",
                                  database="new_brc5")
    connection.autocommit = True

    # Create a cursor to perform database operations
    cursor = connection.cursor()

    wb = Workbook()

    designations_doublees(wb, cursor)
    noms_pas_conformes(wb, cursor)
    commentaires_delateurs(wb, cursor)

    sheet_error(wb, cursor, "20_doublons_post_medium_num.sql", "Même nom post numéro")
    sheet_error(wb, cursor, "30_repetition_ingredients.sql", "Mêmes ingrédients")
    sheet_error(wb, cursor, "35_repetition_ingredients_avec_quantites_et_unites.sql", "Mêmes quantités et unités")
    sheet_error(wb, cursor, "40_meme_commentaire.sql", "Même référence équivalente")

    del wb["Sheet"]
    wb.save(str(path))

create_excel(r"C:\Users\Public\Documents\analyse_redondance_milieux.xlsx")
